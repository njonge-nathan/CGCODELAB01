import openpyxl
import re

# Open the Excel file
excel_file = "Test_Files.xlsx"
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Define a function to generate email addresses
def generate_email(student_name):
    # Remove special characters, spaces, and convert to lowercase
    cleaned_name = re.sub(r'[^a-zA-Z ]', '', student_name).lower()
    # Split the name into parts
    name_parts = cleaned_name.split()
    if len(name_parts) >= 2:
        # If there are at least two parts, use the first letter of the first name and the full last name
        email = f"{name_parts[0][0]}{name_parts[-1]}@gmail.com"
    else:
        # If there's only one name part, use the whole name
        email = f"{name_parts[0]}@gmail.com"
    return email

# Iterate through the Excel rows, generate email addresses, and write them to the fifth column (column E)
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    student_name = row[2]  # Assuming the student name is in the third column (index 2)
    if student_name:
        email = generate_email(student_name)
        sheet.cell(row=row_index, column=5, value=email)  # Save email in the fifth column (column E)

# Save the workbook with the updated email addresses
workbook.save("output_data.xlsx")
